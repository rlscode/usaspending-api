import logging
import os.path

from django.conf import settings
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.db import transaction
import re
import glob

from usaspending_api.references.models import NAICS


class Command(BaseCommand):
    help = "Updates DB from Excel spreadsheets of USAspending terminology definitions into the naics model"

    logger = logging.getLogger('console')

    path = 'usaspending_api/data/naics_archive'
    path = os.path.normpath(path)
    default_path = os.path.join(settings.BASE_DIR, path)

    def add_arguments(self, parser):
        parser.add_argument('-p', '--path', help='the path to the Excel spreadsheets to load', default=self.default_path)
        parser.add_argument('-a', '--append', help='Append to existing guide', action='store_true', default=True)

    def handle(self, *args, **options):
        load_naics(path=options['path'], append=options['append'])


@transaction.atomic
def load_naics(path, append):
    logger = logging.getLogger('console')

    if append:
        logger.info('Appending definitions to existing guide')
    else:
        logger.info('Deleting existing definitions from guide')
        NAICS.objects.all().delete()

    # year regex object precompile
    p_year = re.compile("(20[0-9]{2})")

    dir_files = glob.glob(path + "/*.xlsx")

    for path in sorted(dir_files, reverse=True):
        wb = load_workbook(filename=path)
        ws = wb.active

        naics_year = p_year.search(ws["A1"].value).group()

        for current_row, row in enumerate(ws.rows):
            if current_row == 0:
                continue

            if not row[0].value:
                break  # Reads file only until a blank line

            naics_code = row[0].value
            naics_desc = row[1].value

            obj, created = NAICS.objects.get_or_create(pk=naics_code)

            if not created:
                if int(naics_year) > int(obj.year):
                    NAICS.objects.filter(pk=naics_code).update(description=naics_desc, year=naics_year)
            else:
                obj.description = naics_desc
                obj.year = naics_year
                obj.save()
